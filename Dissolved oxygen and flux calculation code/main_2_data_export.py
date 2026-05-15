import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, borders

# 原始文件与目标文件路径
source_file = "溶解氧及通量结果导出(含hs).xlsx"
output_file = "最终DO浓度提取结果.xlsx"

# 创建一个新的工作簿
wb = openpyxl.Workbook()
wb.remove(wb.active)  # 删除默认的空sheet

# 预设提取要求的时间序列 (0 到 300，间隔20，共16个点)
time_values = list(range(0, 301, 20))

# 预设所需提取列的索引（由于第一行为表头不计入，A列=0, D列=3, K列=10, R列=17）
col_indices = {
    1: 3,  # 1号点位 D列 最终DO浓度
    2: 10,  # 2号点位 K列 最终DO浓度
    3: 17  # 3号点位 R列 最终DO浓度
}

# 设置Excel单元格样式
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center')
thin_border = borders.Border(
    left=borders.Side(style='thin'),
    right=borders.Side(style='thin'),
    top=borders.Side(style='thin'),
    bottom=borders.Side(style='thin')
)

# 循环遍历 6 个工况
for scenario in range(1, 7):
    sheet_name_in_source = f"工况{scenario}"

    try:
        # header=None配合skiprows跳过前两行表头，让索引0直接对应Excel的第3行
        df = pd.read_excel(source_file, sheet_name=sheet_name_in_source, header=None, skiprows=2)
    except Exception as e:
        print(f"读取 {sheet_name_in_source} 时出错或不存在: {e}")
        continue

    # 循环遍历 3 个点位
    for point, col_idx in col_indices.items():
        # 根据命名规则生成新的Sheet名字：如 CS1-S1
        new_sheet_name = f"CS{point}-S{scenario}"
        ws = wb.create_sheet(title=new_sheet_name)

        # 提取数据：从索引0(即第3行)开始，每隔4行(原表每隔3行即步长为4)取到索引60(即第63行)
        extracted_data = df.iloc[0:61:4, col_idx].tolist()

        # 写入 A1：表格标题
        ws['A1'] = new_sheet_name
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = center_align
        ws.merge_cells('A1:B1')  # 合并A1和B1美化
        ws['A1'].border = thin_border
        ws['B1'].border = thin_border

        # 写入 A2, B2：表头
        ws['A2'] = "时间"
        ws['B2'] = "溶解氧"
        ws['A2'].font = bold_font
        ws['B2'].font = bold_font
        ws['A2'].fill = header_fill
        ws['B2'].fill = header_fill
        ws['A2'].alignment = center_align
        ws['B2'].alignment = center_align
        ws['A2'].border = thin_border
        ws['B2'].border = thin_border

        # 写入数据 A3:A18 和 B3:B18
        for i, (t, do) in enumerate(zip(time_values, extracted_data)):
            c_time = ws.cell(row=3 + i, column=1, value=t)
            c_do = ws.cell(row=3 + i, column=2, value=do)

            c_time.alignment = center_align
            c_do.alignment = center_align
            c_time.border = thin_border
            c_do.border = thin_border

            # 对溶解氧数值保留小数展示
            c_do.number_format = '0.00'

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18

# 保存文件
wb.save(output_file)
print(f"数据提取成功！所有点位已导出至：{output_file}")